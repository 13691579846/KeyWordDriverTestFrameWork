"""
------------------------------------
@Time : 2019/8/3 14:20
@Auth : linux超
@File : ParseExcel.py
@IDE  : PyCharm
@Motto: Real warriors,dare to face the bleak warning,dare to face the incisive error!
------------------------------------
"""
from openpyxl import load_workbook
from datetime import datetime

from config.VarConfig import (
    excelPath,
    testStep_testRunTime,
    testStep_testResult,
    testStep_testErrorInfo,
    testStep_testErrorPic
)


class ParseExcel(object):
    """解析excel文件的封装"""

    def __init__(self):
        # 加载excel文件到内存
        self.wb = load_workbook(excelPath)

    def get_row_value(self, sheet_name, raw_no):
        """获取某一行的数据"""
        sh = self.wb[sheet_name]
        row_value_list = []
        for y in range(2, sh.max_column + 1):
            value = sh.cell(raw_no, y).value
            row_value_list.append(value)
        return row_value_list

    def get_column_value(self, sheet_name, col_no):
        """获取某一列的数据"""
        sh = self.wb[sheet_name]
        col_value_list = []
        for x in range(2, sh.max_row + 1):
            value = sh.cell(x, col_no).value
            col_value_list.append(value)
        return col_value_list

    def get_cell_of_value(self, sheet_name, raw_no, col_no):
        """获取某一个单元格的数据"""
        sh = self.wb[sheet_name]
        value = sh.cell(raw_no, col_no).value
        return value

    def write_cell(self, sheet_name, raw_no, col_no, value):
        """向某个单元格写入数据"""
        sh = self.wb[sheet_name]
        sh.cell(raw_no, col_no).value = value
        self.wb.save(excelPath)

    def write_current_time(self, sheet_name, raw_no, col_no):
        """向某个单元格写入当前时间"""
        sh = self.wb[sheet_name]
        time = datetime.now()
        current_time = time.strftime('%Y:%m:%d %H:%M:%S')
        sh.cell(raw_no, col_no).value = current_time
        self.wb.save(excelPath)

    def write_test_result(self, sheet_name, raw_no, result, error_info=None, error_pic=None):
        self.write_current_time(sheet_name, raw_no, testStep_testRunTime)
        self.write_cell(sheet_name, raw_no, testStep_testResult, result)
        if error_info and error_pic:
            self.write_cell(sheet_name, raw_no, testStep_testErrorInfo, error_info)
            self.write_cell(sheet_name, raw_no, testStep_testErrorPic, error_pic)
        else:
            self.write_cell(sheet_name, raw_no, testStep_testErrorInfo, '')
            self.write_cell(sheet_name, raw_no, testStep_testErrorPic, '')


if __name__ == '__main__':
    p = ParseExcel()
    print(p.get_row_value('126account', 2))
    print(p.get_column_value('126account', 3))
    print(p.get_cell_of_value('126account', 2, 3))
